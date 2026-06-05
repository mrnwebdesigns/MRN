#!/usr/bin/env python3
"""Build an MRN builder layout/repeater audit workbook from the local platform runtime."""

from __future__ import annotations

import json
import os
import pathlib
import subprocess
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = pathlib.Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "outputs" / "layout-repeater-audit-20260605"
OUT_FILE = OUT_DIR / "mrn-layout-repeater-audit.xlsx"
WP_PATH = pathlib.Path(
    os.environ.get(
        "MRN_PLATFORM_PATH",
        "/Users/khofmeyer/Development/MRN-sites/platform/public",
    )
)


PHP = r'''
$fields = array(
    "field_mrn_page_hero_rows" => "Hero",
    "field_mrn_page_content_rows" => "Content",
    "field_mrn_page_after_content_rows" => "After Content",
);

function mrn_repeater_audit_field_rows($fields, $path, &$repeaters, &$flexibles, $depth = 0) {
    if (!is_array($fields)) {
        return;
    }

    foreach ($fields as $field) {
        if (!is_array($field)) {
            continue;
        }

        $label = isset($field["label"]) ? (string) $field["label"] : "";
        $name = isset($field["name"]) ? (string) $field["name"] : "";
        $key = isset($field["key"]) ? (string) $field["key"] : "";
        $type = isset($field["type"]) ? (string) $field["type"] : "";
        $field_path = array_merge($path, array($label !== "" ? $label : ($name !== "" ? $name : $key)));

        if ("repeater" === $type) {
            $repeaters[] = array(
                "path" => implode(" > ", array_filter($field_path)),
                "key" => $key,
                "name" => $name,
                "label" => $label,
                "depth" => $depth,
                "subfield_count" => isset($field["sub_fields"]) && is_array($field["sub_fields"]) ? count($field["sub_fields"]) : 0,
            );
        }

        if ("flexible_content" === $type) {
            $layout_count = isset($field["layouts"]) && is_array($field["layouts"]) ? count($field["layouts"]) : 0;
            $flexibles[] = array(
                "path" => implode(" > ", array_filter($field_path)),
                "key" => $key,
                "name" => $name,
                "label" => $label,
                "depth" => $depth,
                "layout_count" => $layout_count,
            );

            if (isset($field["layouts"]) && is_array($field["layouts"])) {
                foreach ($field["layouts"] as $layout) {
                    if (!is_array($layout)) {
                        continue;
                    }

                    $layout_label = isset($layout["label"]) ? (string) $layout["label"] : "";
                    $layout_name = isset($layout["name"]) ? (string) $layout["name"] : "";
                    $layout_path = array_merge($field_path, array($layout_label !== "" ? $layout_label : $layout_name));
                    mrn_repeater_audit_field_rows(isset($layout["sub_fields"]) ? $layout["sub_fields"] : array(), $layout_path, $repeaters, $flexibles, $depth + 1);
                }
            }
        }

        if (isset($field["sub_fields"]) && is_array($field["sub_fields"])) {
            mrn_repeater_audit_field_rows($field["sub_fields"], $field_path, $repeaters, $flexibles, $depth + 1);
        }
    }
}

$result = array(
    "generated_at" => gmdate("c"),
    "builders" => array(),
);

foreach ($fields as $field_key => $builder_label) {
    $field = function_exists("acf_get_field") ? acf_get_field($field_key) : null;
    $layouts = is_array($field) && isset($field["layouts"]) && is_array($field["layouts"]) ? $field["layouts"] : array();
    $builder = array(
        "key" => $field_key,
        "label" => $builder_label,
        "layouts" => array(),
    );

    foreach ($layouts as $layout) {
        if (!is_array($layout)) {
            continue;
        }

        $repeaters = array();
        $flexibles = array();
        $layout_name = isset($layout["name"]) ? (string) $layout["name"] : "";
        $layout_label = isset($layout["label"]) ? (string) $layout["label"] : "";
        mrn_repeater_audit_field_rows(isset($layout["sub_fields"]) ? $layout["sub_fields"] : array(), array($layout_label !== "" ? $layout_label : $layout_name), $repeaters, $flexibles);

        $builder["layouts"][] = array(
            "key" => isset($layout["key"]) ? (string) $layout["key"] : "",
            "name" => $layout_name,
            "label" => $layout_label,
            "repeater_count" => count($repeaters),
            "flexible_count" => count($flexibles),
            "repeaters" => $repeaters,
            "flexibles" => $flexibles,
        );
    }

    $result["builders"][] = $builder;
}

echo wp_json_encode($result);
'''


def fetch_audit_json() -> dict[str, Any]:
    command = ["wp", f"--path={WP_PATH}", "eval", PHP]
    completed = subprocess.run(
        command,
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    return json.loads(completed.stdout)


def append_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 72)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def main() -> None:
    audit = fetch_audit_json()
    summary_rows: list[list[Any]] = []
    repeater_rows: list[list[Any]] = []
    flexible_rows: list[list[Any]] = []

    for builder in audit["builders"]:
        for layout in builder["layouts"]:
            summary_rows.append(
                [
                    builder["label"],
                    builder["key"],
                    layout["name"],
                    layout["label"],
                    layout["key"],
                    layout["repeater_count"],
                    layout["flexible_count"],
                    "\n".join(item["path"] for item in layout["repeaters"]),
                    "\n".join(item["path"] for item in layout["flexibles"]),
                ]
            )

            for repeater in layout["repeaters"]:
                repeater_rows.append(
                    [
                        builder["label"],
                        layout["name"],
                        layout["label"],
                        repeater["path"],
                        repeater["key"],
                        repeater["name"],
                        repeater["label"],
                        repeater["depth"],
                        repeater["subfield_count"],
                    ]
                )

            for flexible in layout["flexibles"]:
                flexible_rows.append(
                    [
                        builder["label"],
                        layout["name"],
                        layout["label"],
                        flexible["path"],
                        flexible["key"],
                        flexible["name"],
                        flexible["label"],
                        flexible["depth"],
                        flexible["layout_count"],
                    ]
                )

    wb = Workbook()
    wb.remove(wb.active)
    append_sheet(
        wb,
        "Layout Summary",
        [
            "Builder",
            "Builder Field Key",
            "Layout Name",
            "Layout Label",
            "Layout Key",
            "Repeater Fields",
            "Nested Flexible Fields",
            "Repeater Paths",
            "Flexible Paths",
        ],
        summary_rows,
    )
    append_sheet(
        wb,
        "Repeater Fields",
        [
            "Builder",
            "Layout Name",
            "Layout Label",
            "Repeater Path",
            "Field Key",
            "Field Name",
            "Field Label",
            "Depth",
            "Subfield Count",
        ],
        repeater_rows,
    )
    append_sheet(
        wb,
        "Flexible Fields",
        [
            "Builder",
            "Layout Name",
            "Layout Label",
            "Flexible Path",
            "Field Key",
            "Field Name",
            "Field Label",
            "Depth",
            "Nested Layout Count",
        ],
        flexible_rows,
    )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    print(OUT_FILE)


if __name__ == "__main__":
    main()
